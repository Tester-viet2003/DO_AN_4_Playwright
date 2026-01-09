# utils/data_loader.py
import os
import csv
import json
from pathlib import Path
from typing import Any, List, Dict, Optional, Tuple, Union
from openpyxl import load_workbook
import pandas as pd

# ============================================================
#  HÀM CHÍNH: load_data() — Đọc đa định dạng cho Data-Driven Test
# ============================================================

def load_data(filename: str) -> List[Tuple]:
    """
    Loader đa năng: đọc dữ liệu từ .md, .csv, .json, .xlsx
    Trả về list[tuple] để dùng cho pytest.mark.parametrize hoặc test DDT.
    
    Args:
        filename (str): Tên file (nằm trong thư mục ../data/)
    
    Returns:
        List[Tuple]: Dữ liệu đọc được
    """
    data: List[Tuple] = []
    base_dir = Path(__file__).resolve().parent.parent / "data"
    file_path = base_dir / filename

    if not file_path.exists():
        raise FileNotFoundError(f" Không tìm thấy file: {file_path}")

    # Markdown (.md)
    if filename.endswith(".md"):
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:  # bỏ 2 dòng đầu (header + separator)
                row = [x.strip() for x in line.strip().split("|") if x.strip()]
                if row:
                    data.append(tuple(row))

    # CSV (.csv)
    elif filename.endswith(".csv"):
        with open(file_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  # bỏ header
            for row in reader:
                if any(row):
                    data.append(tuple(row))

    # JSON (.json)
    elif filename.endswith(".json"):
        with open(file_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)
            if isinstance(json_data, list):
                for item in json_data:
                    data.append(tuple(item.values()))
            elif isinstance(json_data, dict):
                data.append(tuple(json_data.values()))

    # Excel (.xlsx)
    elif filename.endswith(".xlsx"):
        wb = load_workbook(file_path)
        sheet = wb.active
        if sheet is None:
            raise ValueError(" Workbook không chứa sheet nào (sheet = None).")
        if sheet.max_row < 2:
            raise ValueError(" File Excel không có dữ liệu (ít hơn 2 dòng).")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                data.append(tuple(row))

    else:
        raise ValueError(" Định dạng file không được hỗ trợ (.md/.csv/.json/.xlsx).")

    # Kiểm tra dữ liệu rỗng
    if not data:
        print(f" Không có dữ liệu hợp lệ trong file {filename}")
    else:
        print(f" Đã load {len(data)} dòng dữ liệu từ {filename}")

    return data  #  luôn return, giúp Pylance hết cảnh báo


# ============================================================
# HÀM PHỤ: load_excel_data() — Dành cho Keyword-Driven Framework
# ============================================================

def load_excel_data(filepath: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Đọc dữ liệu từ Excel (.xlsx) cho Framework Keyword-Driven.
    Trả về DataFrame chứa cột như TS_ID, Action Keyword, Locator, ...
    """
    file_path_obj = Path(filepath)
    if not file_path_obj.is_file():
        raise FileNotFoundError(f" Không tìm thấy file Excel: {filepath}")

    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name or 0, header=0)
        if df.empty:
            raise ValueError(" File Excel rỗng.")
        print(f" Đã load {len(df)} dòng và {len(df.columns)} cột từ {filepath}")
        return df
    except Exception as e:
        raise Exception(f" Lỗi đọc Excel: {e}")


# ============================================================
# HÀM PHỤ: load_json_config() — Đọc file cấu hình
# ============================================================

def load_json_config(filepath: str) -> Dict[str, Any]:
    """
    Đọc file JSON cấu hình và trả về dict.
    """
    file_path = Path(filepath)
    if not file_path.is_file():
        raise FileNotFoundError(f" Không tìm thấy file cấu hình: {filepath}")

    with open(file_path, encoding="utf-8") as f:
        data = json.load(f)
    print(f" Đã load config từ {filepath}")
    return data


# ============================================================
# DEMO (chạy thử nếu file được chạy trực tiếp)
# ============================================================

if __name__ == "__main__":
    print("\n Demo Data-Driven (.csv/.json/.xlsx/.md)")
    demo_data = load_data("login_data.xlsx")  # đổi sang file của bạn
    print(demo_data[:2])  # in thử 2 dòng đầu

    print("\n Demo Keyword-Driven (.xlsx)")
    df = load_excel_data("data/login_keywords.xlsx")
    print(df.head())
